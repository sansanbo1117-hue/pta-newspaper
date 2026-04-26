#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PTA新聞 Web アプリケーション
Flask ベース、Excel → PDF 自動生成システム
"""

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from pathlib import Path
import os
import json
from datetime import datetime
import tempfile
import traceback
from openpyxl import load_workbook
from weasyprint import HTML
import io

# ─── Flask アプリケーション設定 ───
app = Flask(__name__,
            template_folder='templates',
            static_folder='static')

# ファイルアップロード設定
UPLOAD_FOLDER = Path(tempfile.gettempdir()) / 'pta_newspaper_uploads'
OUTPUT_FOLDER = Path(tempfile.gettempdir()) / 'pta_newspaper_output'
ALLOWED_EXTENSIONS = {'xlsx', 'xlsm'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
app.config['OUTPUT_FOLDER'] = str(OUTPUT_FOLDER)
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# グローバル変数：最新のPDF情報
latest_pdf = {
    'path': None,
    'filename': None
}


# ─── ユーティリティ関数 ───
def allowed_file(filename):
    """ファイル拡張子をチェック"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def load_config(ws):
    """設定シートからデータを読み込み"""
    config = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            config[row[0].strip()] = row[1]
    return config


def load_sections(ws):
    """セクション定義を読み込み"""
    sections = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            section_id = row[1].strip()
            sections[section_id] = {
                "order": row[0],
                "id": section_id,
                "name": row[2] if row[2] else "",
                "en": row[3] if row[3] else "",
                "color": row[4] if row[4] else "",
            }
    return sections


def load_staff(ws):
    """教職員データを読み込み"""
    staff = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            full_name = row[3] if row[3] else "？？"
            role_title = row[2] if row[2] else ""
            furigana = row[4] if row[4] else ""

            staff.append({
                "id": int(row[0]),
                "section": row[1].strip(),
                "role_title": role_title,
                "full_name": full_name,
                "furigana": furigana,
                "message": row[5] if row[5] else "メッセージなし",
                "q1": row[6] if row[6] else "未記入",
                "q2": row[7] if row[7] else "未記入",
            })
    return sorted(staff, key=lambda x: x["id"])


def calculate_grid_layout(total_staff):
    """教職員数から最適なグリッド配置を計算"""
    if total_staff <= 6:
        return 6
    elif total_staff <= 12:
        return 6
    elif total_staff <= 24:
        return 6
    else:
        return min(7, max(5, total_staff // 8))


def get_soft_color(color):
    """メインカラーからソフトカラーを取得"""
    color_map = {
        "bea15a": "#f0e5c8",
        "d47e59": "#f4dfd1",
        "7e9ab8": "#dfe9f2",
        "7f9d85": "#e1ecdf",
        "c88ea2": "#f0dfe6",
    }
    return color_map.get(color.lower(), "#f0e5c8")


def generate_person_card(staff_item, section_color, staff_num):
    """教職員カードのHTMLを生成"""
    role_title_html = ""
    if staff_item.get('role_title'):
        role_title_html = f'<div class="role-title">{staff_item["role_title"]}</div>'

    furigana_html = ""
    if staff_item.get('furigana'):
        furigana_html = f'<div class="furigana">{staff_item["furigana"]}</div>'

    card_html = f'''    <article class="person-card" style="--section-color:#{section_color};--photo:#e2c3aa;">
      <div class="portrait">PHOTO</div>
      <div class="name-row">
        <div class="teacher-info">
          <div class="teacher-name">{staff_item['full_name']}</div>
          {furigana_html}
          {role_title_html}
        </div>
        <div class="no">{staff_num:02d}</div>
      </div>
      <div class="message">{staff_item['message']}</div>
      <div class="tags">
        <div class="tag"><b>Q1</b><span>{staff_item['q1']}</span></div>
        <div class="tag"><b>Q2</b><span>{staff_item['q2']}</span></div>
      </div>
    </article>
'''
    return card_html


def generate_section_html(section, sections_def, staff_all):
    """セクションのHTMLを生成"""
    sec_info = sections_def.get(section, {})
    color = sec_info.get("color", "bea15a").lstrip("#")

    section_staff = [s for s in staff_all if s["section"] == section]
    section_staff = sorted(section_staff, key=lambda x: x["id"])

    if not section_staff:
        return ""

    cols = calculate_grid_layout(len(section_staff))

    grid_html = f'''  <div class="section" style="--section-color:#{color};--section-soft:{get_soft_color(color)};">
    <div class="section-head">
      <div class="section-name">{sec_info.get('name', section)}</div>
      <div class="section-meta">
        <div class="section-count">{len(section_staff)}</div>
        <div class="section-en">{sec_info.get('en', '').upper()}</div>
      </div>
    </div>
    <div class="section-body">
      <div class="people-grid" style="--cols:{cols};">
'''

    for staff_item in section_staff:
        card = generate_person_card(staff_item, color, staff_item["id"])
        grid_html += card

    grid_html += '''      </div>
    </div>
  </div>
'''
    return grid_html


def generate_legend_html(sections_def, staff_all):
    """凡例HTMLを生成"""
    legend = '<span><i class="dot" style="--dot:#bea15a"></i>管理職・専門</span>'
    legend += '<span><i class="dot" style="--dot:#d47e59"></i>小学部</span>'
    legend += '<span><i class="dot" style="--dot:#7e9ab8"></i>中学部</span>'
    legend += '<span><i class="dot" style="--dot:#7f9d85"></i>高等部</span>'
    legend += '<span><i class="dot" style="--dot:#c88ea2"></i>事務・支援</span>'
    return legend


def process_excel_file(file_path):
    """
    Excel ファイルを処理して PDF を生成する
    """
    global latest_pdf

    try:
        # Excelファイルを読み込む
        wb = load_workbook(file_path)
        config = load_config(wb["設定"])
        sections_def = load_sections(wb["セクション定義"])
        staff_all = load_staff(wb["教職員"])

        # テンプレートを読み込む
        template_path = Path(__file__).parent / "html_template.html"
        if not template_path.exists():
            raise FileNotFoundError(f"html_template.html が見つかりません: {template_path}")

        with open(template_path, 'r', encoding='utf-8') as f:
            html = f.read()

        # 設定値を取得
        year = config.get("年度", "2026")
        vol = config.get("号数", "1")
        school = config.get("学校名", "学校")
        paper_name = config.get("新聞名", "陽だまり")
        title = config.get("タイトル", "はじめまして、今年の先生たち。")
        description = config.get("副タイトル", "")
        year_ja = str(int(year) - 2018)

        # プレースホルダーを置換
        html = html.replace("{{YEAR}}", str(year))
        html = html.replace("{{VOL}}", str(vol))
        html = html.replace("{{SCHOOL}}", str(school))
        html = html.replace("{{TITLE}}", str(title))
        html = html.replace("{{PAPER_NAME}}", str(paper_name))
        html = html.replace("{{YEAR_JA}}", str(year_ja))
        html = html.replace("{{DESCRIPTION}}", str(description))
        html = html.replace("{{Q1}}", "お休みの日の楽しみや過ごし方を教えてください。")
        html = html.replace("{{Q2}}", "子どものころの思い出や、どんなお子さんだったか教えてください。")

        # 教職員セクションを生成
        section_order = sorted(sections_def.items(), key=lambda x: x[1].get("order", 99))
        staff_html = ""
        for section_id, sec_info in section_order:
            section_html = generate_section_html(section_id, sections_def, staff_all)
            staff_html += section_html

        html = html.replace('<section class="staff-atlas" id="staffAtlas"></section>',
                           f'<section class="staff-atlas" id="staffAtlas">\n{staff_html}  </section>')

        # 凡例を生成
        legend_html = generate_legend_html(sections_def, staff_all)
        html = html.replace('<div class="legend" id="legend"></div>',
                           f'<div class="legend" id="legend">{legend_html}</div>')

        # PDFファイル名を生成
        pdf_filename = f"PTA新聞_{paper_name}_{year}年度_第{vol}号.pdf"
        pdf_path = Path(app.config['OUTPUT_FOLDER']) / pdf_filename

        # HTMLからPDFを生成
        HTML(string=html).write_pdf(str(pdf_path))

        # グローバル変数を更新
        latest_pdf['path'] = str(pdf_path)
        latest_pdf['filename'] = pdf_filename

        return str(pdf_path), pdf_filename

    except Exception as e:
        print(f"Error processing Excel: {e}")
        print(traceback.format_exc())
        raise


# ─── ルーティング ───

@app.route('/')
def index():
    """メインページ"""
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Excel ファイルアップロード + 処理"""
    try:
        # ファイル確認
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'ファイルが選択されていません'
            }), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'ファイルが選択されていません'
            }), 400

        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': '対応しているファイル形式は .xlsx または .xlsm です'
            }), 400

        # ファイル保存（セキュアなファイル名生成）
        # 元のファイル名から拡張子を抽出（日本語ファイル名対応）
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'xlsx'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # タイムスタンプベースのセキュアなファイル名を使用
        unique_filename = f"{timestamp}.{ext}"
        filename = file.filename  # 元のファイル名を保持
        file_path = Path(app.config['UPLOAD_FOLDER']) / unique_filename

        file.save(str(file_path))

        # ─── PDF 生成処理 ───
        pdf_path, pdf_filename = process_excel_file(file_path)

        return jsonify({
            'success': True,
            'message': 'PDFが生成されました',
            'filename': filename,
            'pdf_filename': pdf_filename,
            'pdfUrl': f'/api/download/{pdf_filename}',
            'upload_time': timestamp
        }), 200

    except Exception as e:
        print(f"Error: {e}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'処理中にエラーが発生しました: {str(e)}'
        }), 500


@app.route('/api/download/<filename>', methods=['GET'])
def download_pdf(filename):
    """PDFファイルをダウンロード"""
    try:
        file_path = Path(app.config['OUTPUT_FOLDER']) / filename

        if not file_path.exists():
            return jsonify({
                'success': False,
                'error': 'ファイルが見つかりません'
            }), 404

        return send_file(
            str(file_path),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error: {e}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'ダウンロード中にエラーが発生しました: {str(e)}'
        }), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """ヘルスチェック"""
    return jsonify({'status': 'ok'}), 200


# ─── エラーハンドリング ───

@app.errorhandler(413)
def too_large(e):
    """ファイルサイズが大きすぎる"""
    return jsonify({
        'success': False,
        'error': 'ファイルサイズが大きすぎます（最大10MB）'
    }), 413


@app.errorhandler(404)
def not_found(e):
    """ページが見つからない"""
    return jsonify({
        'success': False,
        'error': 'ページが見つかりません'
    }), 404


@app.errorhandler(500)
def internal_error(e):
    """サーバーエラー"""
    return jsonify({
        'success': False,
        'error': 'サーバーエラーが発生しました'
    }), 500




# ─── メイン ───

if __name__ == '__main__':
    # 開発環境
    app.run(debug=True, host='0.0.0.0', port=5000)

    # 本番環境（Render など）では PORT 環境変数を使用
    # port = int(os.environ.get('PORT', 5000))
    # app.run(debug=False, host='0.0.0.0', port=port)
